# -*- coding: utf-8 -*-
# @File    : do_excel.PY
# @Date    : 2019/7/8-17:19
# @Author  : leihuijuan
# @Emali   : huijuan_lei@163.com

from openpyxl import load_workbook

class Case_Excel:
    def __init__(self):
        self.case_id=None
        self.title=None
        self.url=None
        self.method=None
        self.data=None
        self.expected=None
        self.actual=None
        self.result=None
        self.check_sql=None

class Do_Excel:
    def __init__(self,filename,index):
        self.filename=filename
        self.wb=load_workbook(self.filename)
        self.sheet=self.wb.worksheets[index]#根据工作薄的索引取哪个工作薄
        self.index=index

    #读excel表格
    def read_excel(self):
        list_data=[]
        for i in range(2,self.sheet.max_row+1):
            case =Case_Excel()  #实例化Case_Excel()---对象实例化调用
            case.case_id=self.sheet.cell(row=i,column=1).value     #case_id
            case.title=self.sheet.cell(row=i,column=2).value     #title
            case.url=self.sheet.cell(row=i,column=3) .value   #url
            case.method=self.sheet.cell(row=i,column=4).value   #method
            case.data=self.sheet.cell(row=i,column=5).value     #data
            case.expected = self.sheet.cell(row=i,column=6).value   #expected
            case.check_sql=self.sheet.cell(row=i,column=9).value    #读验证sql语句
            list_data.append(case)
        self.wb.close()
        return list_data

    def write_excel(self,row,actual,result):
        sheet=self.wb.worksheets[self.index]
        sheet.cell(row=row,column=7).value=actual
        sheet.cell(row=row,column=8).value=result
        self.wb.save(filename=self.filename)
        self.wb.close()

if __name__ == '__main__':
    excel=Do_Excel("../test_data/Future_loans.xlsx",0)
    res=excel.read_excel()
    for i in res:
        print(i.__dict__)



